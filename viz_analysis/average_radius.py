#!/usr/bin/python3
#########################################################################
#    SPF - Stochastic Phase Field
#    Copyright (C) 2025 
#    Peng Geng <penggeng@g.ucla.edu>
#
#    This program is free software; you can redistribute it and/or modify
#    it under the terms of the GNU General Public License as published by
#    the Free Software Foundation; either version 2 of the License, or
#    (at your option) any later version.
#
#    This program is distributed in the hope that it will be useful,
#    but WITHOUT ANY WARRANTY; without even the implied warranty of
#    MERCHANTABILITY or FITNESS FOR A PARTICULAR PURPOSE.  See the
#    GNU General Public License for more details.
#
#    You should have received a copy of the GNU General Public License along
#    with this program; if not, write to the Free Software Foundation, Inc.,
#    51 Franklin Street, Fifth Floor, Boston, MA 02110-1301 USA.
#
#    See the README file in the top-level SPF directory.
#########################################################################
# File: average_radius.py
# Purpose: This script estimates the average cluster radius from the
# simulation output HDF5 files by applying a threshold to identify 
# spherical cluster voxels.
# Requirements: numpy, h5py, vtk, scipy, openpyxl


import h5py
import vtk
import numpy as np
from scipy.ndimage import label
import openpyxl


# Specify the HDF5 file path
file_path = "50.h5"

# Specify the steps to convert
steps = range(0, 100001, 10000)

# Create a new workbook
workbook = openpyxl.Workbook()

# Get the active worksheet
worksheet = workbook.active

# Create lists to store results(sheet)
result_step = []
result_mean_equivalent_radius = []

# Convert each step to VTI file
for step in steps:
    with h5py.File(file_path, "r") as f:
        step_name = "Step{}".format(step)
        if step_name in f:
            # Access the dataset
            dataset = f[step_name + "/phi"]
            
            # Convert the dataset to a NumPy array
            data = np.array(dataset).astype(np.float32)
            
            """
            # Nucleation ///////////////////////////////////////////
            # Aisosurface threshold > 0.5
            data1 = data > 0.5
           
            # Label connected components
            labeled_array, num_features = label(data1)
           
            # Measure the size of each connected component
            component_sizes = np.bincount(labeled_array.ravel())
           
            # Filter out small components based on a size threshold
            size_threshold = 4 ** 3  # Adjust threshold as needed
            remove_indices = np.where(component_sizes < size_threshold)[0]
           
            # Create a filtered binary array
            for index in remove_indices:
                data[labeled_array == index] = 0
            # Nucleation ///////////////////////////////////////////
            """
            
            # Aisosurface threshold > 0.5
            data1 = data > 0.5
            
            # Label connected components
            labeled_array, num_features = label(data1)
            
            # Measure the size of each connected component
            component_sizes = np.bincount(labeled_array.ravel())
            
            # Ignore the background label (index 0)
            component_sizes = component_sizes[1:]

            # Compute equivalent radius for each domain
            equivalent_radii = (3 * component_sizes / (4 * np.pi)) ** (1/3)

            # Compute mean equivalent radius
            mean_equivalent_radius = np.mean(equivalent_radii)

            # Print the result
            #print("Mean Equivalent Radius:", step, mean_equivalent_radius)    
            
            # Add the unique k value and its maximum C value to results
            result_step.append(step)
            result_mean_equivalent_radius.append(mean_equivalent_radius)

# Write the headers
worksheet.cell(row=1, column=1, value="Step")
worksheet.cell(row=1, column=2, value="mean_equivalent_radius")

for row, (Step, mean_equivalent_radius) in enumerate(zip(result_step, result_mean_equivalent_radius), start=2):
    worksheet.cell(row=row, column=1, value=Step)
    worksheet.cell(row=row, column=2, value=mean_equivalent_radius)

# Save the workbook to a file
workbook.save("mean_equivalent_radius.xlsx")



            
            
            
            
            
            
 
            
