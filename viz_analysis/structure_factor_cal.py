#!/usr/bin/python3
#########################################################################
#    SPF - Stochastic Phase Field
#    Copyright (C) 2025 
#    Peng Geng <penggeng@g.ucla.edu>
#
#    This program is free software; you can redistribute it and/or modify
#    it under the terms of the GNU General Public License as published by
#    the Free Software Foundation; either version 2 of the License, or
#    (at your option) any later version.
#
#    This program is distributed in the hope that it will be useful,
#    but WITHOUT ANY WARRANTY; without even the implied warranty of
#    MERCHANTABILITY or FITNESS FOR A PARTICULAR PURPOSE.  See the
#    GNU General Public License for more details.
#
#    You should have received a copy of the GNU General Public License along
#    with this program; if not, write to the Free Software Foundation, Inc.,
#    51 Franklin Street, Fifth Floor, Boston, MA 02110-1301 USA.
#
#    See the README file in the top-level SPF directory.
#########################################################################
# File: structure_factor_cal.py
# Purpose: This script computes the structure factor from the simulation
# output HDF5 file using the method described in Section 2.2 of the
# referenced paper.
# Requirements: numpy, h5py, matplotlib, scipy, openpyxl


import h5py
import numpy as np

import matplotlib.pyplot as plt
from scipy.signal import find_peaks
import openpyxl

from collections import defaultdict

# Specify the HDF5 file path
file_path = "50.h5"

# Specify the steps to convert
steps = range(10000, 100001, 10000)

# Create a new workbook
workbook = openpyxl.Workbook()

# Get the active worksheet
worksheet = workbook.active

i=1
# Convert each step to VTI file
for step in steps:
    with h5py.File(file_path, "r") as f:
        step_name = "Step{}".format(step)
        #step_name = "Step{0}"
        if step_name in f:
            # Access the dataset
            dataset = f[step_name + "/phi"]
            
            # Convert the dataset to a NumPy array
            c_real = np.array(dataset)
            c_real = c_real - np.mean(c_real)
            
            # Perform FFT
            C_k = np.fft.fftn(c_real)
            C_k_mag_squared = np.abs(C_k)**2
    
            # Normalize by the total number of voxels
            C_k_norm = C_k_mag_squared / np.prod(c_real.shape)
    
            # Get grid dimensions and wavevector components
            nx, ny, nz = c_real.shape
            kx = np.fft.fftfreq(nx, d=1) * 2 * np.pi
            ky = np.fft.fftfreq(ny, d=1) * 2 * np.pi
            kz = np.fft.fftfreq(nz, d=1) * 2 * np.pi
            k_grid = np.sqrt(kx[:, None, None]**2 + ky[None, :, None]**2 + kz[None, None, :]**2)
    
            # Flatten arrays
            k_flat = k_grid.flatten()
            C_k_flat = C_k_norm.flatten()
            
            # Round to four decimal
            #k_flat_r = np.round(k_flat, 4)
            #C_k_flat_r = np.round(C_k_flat, 4)
        
            # Create lists to store results
            result_k = []
            result_C = []
            
            # Find unique values in k_flat
            unique_k = np.unique(k_flat)

            # For each unique value, find the maximum C value
            for k in unique_k:
                # Find indices where k_flat matches the current unique value
                indices = np.where(k_flat == k)[0]
    
                # Get the corresponding C values and average
                avg_c = np.mean(C_k_flat[indices])
    
                # Add the unique k value and its maximum C value to results
                result_k.append(k)
                result_C.append(avg_c)

            # Write the headers
            worksheet.cell(row=1, column=i, value=f"Step{step}")
            worksheet.cell(row=1, column=i+1, value=f"Step{step}")
            worksheet.cell(row=2, column=i, value="k")
            worksheet.cell(row=2, column=i+1, value="c")


            for row, (k, c) in enumerate(zip(result_k, result_C), start=3):
                worksheet.cell(row=row, column=i, value=k)
                worksheet.cell(row=row, column=i+1, value=c)

            i +=3
            
            #Peak detect
            #result_k = np.array(result_k)
            #result_C = np.array(result_C)
            # Detect peaks
            #peaks, properties = find_peaks(result_C, height=0.1, distance=1)  # Adjust `height` and `distance`
            
            # Extract peak positions and values
            #peak_k = result_k[peaks]
            #peak_S = result_C[peaks]
            
            # Plot S(k)
            #plt.scatter(k_flat, C_k_flat, s=15, label="Original")
            plt.scatter(result_k, result_C, s=15, label="average")
            #plt.scatter(peak_k, peak_S, s=5, label="Detected Peaks")
            plt.xlabel(r"$k$")
            plt.ylabel(r"$S(k)$")
            plt.title("Structure Factor")
            plt.grid()
            plt.legend()
            plt.show()    

# Save the workbook to a file
workbook.save("Structure_factor.xlsx")







